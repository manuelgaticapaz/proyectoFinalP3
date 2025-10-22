from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Q, Avg
from datetime import datetime, timedelta
import json

# Importaciones para reportes (con manejo de errores)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')  # Para usar matplotlib sin GUI
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from io import BytesIO
import base64

from appointments.models import Appointment
from doctors.models import Doctor
from patients.models import Patient
from clinicas.models import Clinica
from .models import ReporteGenerado

@login_required
def dashboard_reportes(request):
    """
    Dashboard principal de reportes
    """
    # Obtener doctor y clínica del usuario
    try:
        doctor = Doctor.objects.get(usuario=request.user)
        clinica = doctor.clinica
    except Doctor.DoesNotExist:
        doctor = None
        clinica = None
    
    # Estadísticas rápidas
    hoy = timezone.now().date()
    inicio_mes = hoy.replace(day=1)
    
    # Filtrar por clínica si existe
    citas_query = Appointment.objects.all()
    if clinica:
        citas_query = citas_query.filter(clinica=clinica)
    
    estadisticas = {
        'citas_hoy': citas_query.filter(fecha__date=hoy).count(),
        'citas_mes': citas_query.filter(fecha__date__gte=inicio_mes).count(),
        'pacientes_activos': Patient.objects.filter(clinica=clinica).count() if clinica else Patient.objects.count(),
        'doctores_activos': Doctor.objects.filter(clinica=clinica, activo=True).count() if clinica else Doctor.objects.filter(activo=True).count(),
    }
    
    # Reportes recientes
    reportes_recientes = ReporteGenerado.objects.filter(
        generado_por=request.user
    ).order_by('-fecha_generacion')[:5]
    
    if clinica:
        reportes_recientes = reportes_recientes.filter(clinica=clinica)
    
    context = {
        'estadisticas': estadisticas,
        'reportes_recientes': reportes_recientes,
        'clinica': clinica,
    }
    
    return render(request, 'reportes/dashboard.html', context)

@login_required
def generar_reporte_pdf(request):
    """
    Generar reporte en formato PDF
    """
    if not REPORTLAB_AVAILABLE:
        return JsonResponse({
            'error': 'ReportLab no está disponible. Por favor instale: pip install reportlab'
        }, status=500)
    
    tipo_reporte = request.GET.get('tipo', 'citas_diarias')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    doctor_id = request.GET.get('doctor_id')
    
    # Validar fechas
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    else:
        fecha_inicio = timezone.now().date()
    
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    else:
        fecha_fin = fecha_inicio
    
    # Obtener clínica del usuario
    try:
        doctor_usuario = Doctor.objects.get(usuario=request.user)
        clinica = doctor_usuario.clinica
    except Doctor.DoesNotExist:
        clinica = None
    
    # Crear respuesta HTTP para PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo_reporte}_{fecha_inicio}.pdf"'
    
    # Crear documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    elementos = []
    estilos = getSampleStyleSheet()
    
    # Título del reporte
    titulo_style = ParagraphStyle(
        'TituloReporte',
        parent=estilos['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Centrado
        textColor=colors.darkblue
    )
    
    titulo = f"Reporte de {tipo_reporte.replace('_', ' ').title()}"
    elementos.append(Paragraph(titulo, titulo_style))
    elementos.append(Spacer(1, 20))
    
    # Información de la clínica
    if clinica:
        info_clinica = f"<b>Clínica:</b> {clinica.nombre}<br/>"
        info_clinica += f"<b>Período:</b> {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        elementos.append(Paragraph(info_clinica, estilos['Normal']))
        elementos.append(Spacer(1, 20))
    
    # Obtener datos según el tipo de reporte
    if tipo_reporte == 'citas_diarias':
        datos = obtener_datos_citas_diarias(fecha_inicio, fecha_fin, doctor_id, clinica)
    elif tipo_reporte == 'estadisticas_doctor':
        datos = obtener_estadisticas_doctor(fecha_inicio, fecha_fin, doctor_id, clinica)
    elif tipo_reporte == 'ocupacion_clinica':
        datos = obtener_ocupacion_clinica(fecha_inicio, fecha_fin, clinica)
    else:
        datos = []
    
    # Crear tabla con los datos
    if datos:
        tabla = Table(datos)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elementos.append(tabla)
    else:
        elementos.append(Paragraph("No se encontraron datos para el período seleccionado.", estilos['Normal']))
    
    # Generar PDF
    doc.build(elementos)
    
    # Guardar registro del reporte
    ReporteGenerado.objects.create(
        nombre=titulo,
        tipo=tipo_reporte,
        formato='pdf',
        filtros_aplicados={
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'doctor_id': doctor_id,
        },
        generado_por=request.user,
        clinica=clinica
    )
    
    return response

@login_required
def generar_reporte_excel(request):
    """
    Generar reporte en formato Excel
    """
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'error': 'OpenPyXL no está disponible. Por favor instale: pip install openpyxl'
        }, status=500)
    
    tipo_reporte = request.GET.get('tipo', 'citas_diarias')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    doctor_id = request.GET.get('doctor_id')
    
    # Validar fechas
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    else:
        fecha_inicio = timezone.now().date()
    
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    else:
        fecha_fin = fecha_inicio
    
    # Obtener clínica del usuario
    try:
        doctor_usuario = Doctor.objects.get(usuario=request.user)
        clinica = doctor_usuario.clinica
    except Doctor.DoesNotExist:
        clinica = None
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo_reporte}"
    
    # Estilos
    titulo_font = Font(size=16, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    titulo_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte de {tipo_reporte.replace('_', ' ').title()}"
    ws['A1'].font = titulo_font
    ws['A1'].fill = titulo_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información de la clínica
    if clinica:
        ws['A2'] = f"Clínica: {clinica.nombre}"
        ws['A3'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    
    # Obtener datos
    if tipo_reporte == 'citas_diarias':
        datos = obtener_datos_citas_excel(fecha_inicio, fecha_fin, doctor_id, clinica)
    elif tipo_reporte == 'estadisticas_doctor':
        datos = obtener_estadisticas_doctor_excel(fecha_inicio, fecha_fin, doctor_id, clinica)
    else:
        datos = []
    
    # Escribir datos
    if datos:
        # Headers
        fila_inicio = 5
        for col, header in enumerate(datos[0], 1):
            celda = ws.cell(row=fila_inicio, column=col, value=header)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')
        
        # Datos
        for fila, registro in enumerate(datos[1:], fila_inicio + 1):
            for col, valor in enumerate(registro, 1):
                ws.cell(row=fila, column=col, value=valor)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo_reporte}_{fecha_inicio}.xlsx"'
    
    wb.save(response)
    
    # Guardar registro del reporte
    ReporteGenerado.objects.create(
        nombre=f"Reporte de {tipo_reporte.replace('_', ' ').title()}",
        tipo=tipo_reporte,
        formato='excel',
        filtros_aplicados={
            'fecha_inicio': fecha_inicio.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'doctor_id': doctor_id,
        },
        generado_por=request.user,
        clinica=clinica
    )
    
    return response

@login_required
def generar_grafico_estadisticas(request):
    """
    Generar gráficos estadísticos
    """
    if not MATPLOTLIB_AVAILABLE:
        return JsonResponse({
            'error': 'Matplotlib no está disponible. Por favor instale: pip install matplotlib'
        }, status=500)
    
    tipo_grafico = request.GET.get('tipo', 'citas_por_mes')
    
    # Obtener clínica del usuario
    try:
        doctor_usuario = Doctor.objects.get(usuario=request.user)
        clinica = doctor_usuario.clinica
    except Doctor.DoesNotExist:
        clinica = None
    
    # Crear gráfico según el tipo
    plt.figure(figsize=(12, 8))
    
    if tipo_grafico == 'citas_por_mes':
        datos = obtener_citas_por_mes(clinica)
        plt.bar(datos['meses'], datos['cantidades'], color='skyblue')
        plt.title('Citas por Mes', fontsize=16, fontweight='bold')
        plt.xlabel('Mes')
        plt.ylabel('Cantidad de Citas')
        
    elif tipo_grafico == 'citas_por_doctor':
        datos = obtener_citas_por_doctor(clinica)
        plt.pie(datos['cantidades'], labels=datos['doctores'], autopct='%1.1f%%')
        plt.title('Distribución de Citas por Doctor', fontsize=16, fontweight='bold')
        
    elif tipo_grafico == 'estados_citas':
        datos = obtener_estados_citas(clinica)
        colores = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', '#ff99cc']
        plt.pie(datos['cantidades'], labels=datos['estados'], autopct='%1.1f%%', colors=colores)
        plt.title('Estados de las Citas', fontsize=16, fontweight='bold')
    
    plt.tight_layout()
    
    # Convertir gráfico a base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode()
    plt.close()
    
    return JsonResponse({
        'grafico': f"data:image/png;base64,{imagen_base64}",
        'tipo': tipo_grafico
    })

# Funciones auxiliares para obtener datos
def obtener_datos_citas_diarias(fecha_inicio, fecha_fin, doctor_id, clinica):
    """Obtener datos de citas diarias para PDF"""
    citas_query = Appointment.objects.filter(
        fecha__date__gte=fecha_inicio,
        fecha__date__lte=fecha_fin
    )
    
    if clinica:
        citas_query = citas_query.filter(clinica=clinica)
    
    if doctor_id:
        citas_query = citas_query.filter(doctor_id=doctor_id)
    
    citas = citas_query.select_related('paciente', 'doctor').order_by('fecha')
    
    datos = [['Fecha', 'Hora', 'Paciente', 'Doctor', 'Estado']]
    
    for cita in citas:
        datos.append([
            cita.fecha.strftime('%d/%m/%Y'),
            cita.fecha.strftime('%H:%M'),
            str(cita.paciente),
            str(cita.doctor),
            cita.get_estado_display()
        ])
    
    return datos

def obtener_datos_citas_excel(fecha_inicio, fecha_fin, doctor_id, clinica):
    """Obtener datos de citas para Excel"""
    return obtener_datos_citas_diarias(fecha_inicio, fecha_fin, doctor_id, clinica)

def obtener_estadisticas_doctor(fecha_inicio, fecha_fin, doctor_id, clinica):
    """Obtener estadísticas por doctor para PDF"""
    doctores_query = Doctor.objects.filter(activo=True)
    
    if clinica:
        doctores_query = doctores_query.filter(clinica=clinica)
    
    if doctor_id:
        doctores_query = doctores_query.filter(id=doctor_id)
    
    datos = [['Doctor', 'Total Citas', 'Completadas', 'Canceladas', 'Porcentaje Éxito']]
    
    for doctor in doctores_query:
        citas_doctor = Appointment.objects.filter(
            doctor=doctor,
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        )
        
        total = citas_doctor.count()
        completadas = citas_doctor.filter(estado='completada').count()
        canceladas = citas_doctor.filter(estado='cancelada').count()
        porcentaje = (completadas / total * 100) if total > 0 else 0
        
        datos.append([
            str(doctor),
            total,
            completadas,
            canceladas,
            f"{porcentaje:.1f}%"
        ])
    
    return datos

def obtener_estadisticas_doctor_excel(fecha_inicio, fecha_fin, doctor_id, clinica):
    """Obtener estadísticas por doctor para Excel"""
    return obtener_estadisticas_doctor(fecha_inicio, fecha_fin, doctor_id, clinica)

def obtener_ocupacion_clinica(fecha_inicio, fecha_fin, clinica):
    """Obtener datos de ocupación de clínica"""
    datos = [['Fecha', 'Citas Programadas', 'Citas Completadas', 'Porcentaje Ocupación']]
    
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        citas_query = Appointment.objects.filter(fecha__date=fecha_actual)
        
        if clinica:
            citas_query = citas_query.filter(clinica=clinica)
        
        programadas = citas_query.count()
        completadas = citas_query.filter(estado='completada').count()
        porcentaje = (completadas / programadas * 100) if programadas > 0 else 0
        
        datos.append([
            fecha_actual.strftime('%d/%m/%Y'),
            programadas,
            completadas,
            f"{porcentaje:.1f}%"
        ])
        
        fecha_actual += timedelta(days=1)
    
    return datos

def obtener_citas_por_mes(clinica):
    """Obtener datos de citas por mes para gráficos"""
    # Últimos 12 meses
    hoy = timezone.now().date()
    hace_12_meses = hoy - timedelta(days=365)
    
    citas_query = Appointment.objects.filter(fecha__date__gte=hace_12_meses)
    
    if clinica:
        citas_query = citas_query.filter(clinica=clinica)
    
    # Agrupar por mes
    meses = []
    cantidades = []
    
    for i in range(12):
        mes_inicio = (hoy.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        if i == 0:
            mes_fin = hoy
        else:
            mes_fin = (mes_inicio.replace(month=mes_inicio.month+1) if mes_inicio.month < 12 
                     else mes_inicio.replace(year=mes_inicio.year+1, month=1)) - timedelta(days=1)
        
        cantidad = citas_query.filter(
            fecha__date__gte=mes_inicio,
            fecha__date__lte=mes_fin
        ).count()
        
        meses.insert(0, mes_inicio.strftime('%b %Y'))
        cantidades.insert(0, cantidad)
    
    return {'meses': meses, 'cantidades': cantidades}

def obtener_citas_por_doctor(clinica):
    """Obtener distribución de citas por doctor"""
    doctores_query = Doctor.objects.filter(activo=True)
    
    if clinica:
        doctores_query = doctores_query.filter(clinica=clinica)
    
    doctores = []
    cantidades = []
    
    for doctor in doctores_query:
        cantidad = Appointment.objects.filter(doctor=doctor).count()
        if cantidad > 0:
            doctores.append(f"Dr. {doctor.nombre} {doctor.apellidos}")
            cantidades.append(cantidad)
    
    return {'doctores': doctores, 'cantidades': cantidades}

def obtener_estados_citas(clinica):
    """Obtener distribución de estados de citas"""
    citas_query = Appointment.objects.all()
    
    if clinica:
        citas_query = citas_query.filter(clinica=clinica)
    
    estados_count = citas_query.values('estado').annotate(
        cantidad=Count('estado')
    ).order_by('-cantidad')
    
    estados = []
    cantidades = []
    
    for item in estados_count:
        estados.append(item['estado'].title())
        cantidades.append(item['cantidad'])
    
    return {'estados': estados, 'cantidades': cantidades}
